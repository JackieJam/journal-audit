"""
Excel 输出模块：生成审计抽样报告。
修复原版问题：
- 按凭证合并（不再每行重复 LLM 判断）
- 严格按 max_sample_size 截断
- 规则类型去重（同凭证命中多规则时合并展示）
- LLM 判断为空的凭证不出现在样本清单
"""

from __future__ import annotations

from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    import pandas as pd
    from modules.llm_verifier import LLMJudgment
    from modules.rule_engine import RuleHit, RuleResult

HEADER_FONT = Font(bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
MED_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

SAMPLE_COLS = [
    ("序号", 6), ("凭证编号", 14), ("组合ID", 24), ("关联凭证", 28), ("年份", 8), ("过账日期", 12), ("凭证类型", 10),
    ("总账科目", 12), ("总账科目名称", 22), ("借/贷", 8), ("金额", 16),
    ("文本", 30), ("供应商", 22), ("客户", 22), ("用户名", 12),
    ("规则类型", 20), ("触发证据", 35), ("组合证据", 50),
    ("LLM风险级别", 12), ("LLM判断理由", 35), ("建议核查程序", 35),
    ("核查结论", 20),  # 留给审计师填写
]


def generate_report(
    df: "pd.DataFrame",
    rule_results: list["RuleResult"],
    llm_judgments: dict[str, list["LLMJudgment"]],
    output_path: str,
    max_sample_size: int = 50,
    manual_final_samples: list[dict] | None = None,
) -> dict:
    """
    生成 Excel 报告，返回统计摘要 dict。
    """
    # ── 构建凭证维度的合并视图 ──
    judgment_lookup = _build_judgment_lookup(llm_judgments)
    hit_lookup = _build_hit_lookup(rule_results)

    if judgment_lookup:
        # 有 LLM 核实：按风险级别排序，取 top N
        confirmed_vids = sorted(
            judgment_lookup.keys(),
            key=lambda v: judgment_lookup[v].risk_level == "高",
            reverse=True,
        )[:max_sample_size]
    else:
        # 无 LLM：按规则命中优先级排序，取 top N 凭证
        voucher_priority: dict[str, int] = {}
        for rr in rule_results:
            for hit in rr.hits:
                vid = hit.voucher_id
                voucher_priority[vid] = max(voucher_priority.get(vid, 0), hit.priority)
        confirmed_vids = sorted(
            voucher_priority.keys(),
            key=lambda v: voucher_priority[v],
            reverse=True,
        )[:max_sample_size]

    wb = Workbook()
    _write_sample_sheet(wb, df, confirmed_vids, hit_lookup, judgment_lookup)
    if manual_final_samples:
        _write_manual_final_sheet(wb, df, manual_final_samples)
    _write_stats_sheet(wb, rule_results, llm_judgments)
    _write_rules_sheet(wb, rule_results)

    wb.save(output_path)

    total_unique_vouchers = len({h.voucher_id for rr in rule_results for h in rr.hits})
    return {
        "output_path": output_path,
        "sample_vouchers": len(confirmed_vids),
        "total_rule_hits": sum(r.count for r in rule_results),
        "total_unique_vouchers": total_unique_vouchers,
        "llm_confirmed": len(judgment_lookup),
        "high_risk": sum(1 for j in judgment_lookup.values() if j.risk_level == "高") if judgment_lookup else 0,
        "medium_risk": sum(1 for j in judgment_lookup.values() if j.risk_level == "中") if judgment_lookup else 0,
        "manual_final_vouchers": len({
            str(vid)
            for group in manual_final_samples or []
            for vid in group.get("voucher_ids", [])
        }),
    }


def _build_judgment_lookup(llm_judgments: dict) -> dict[str, "LLMJudgment"]:
    lookup: dict = {}
    for judgments in llm_judgments.values():
        for j in judgments:
            if j.confirmed and j.voucher_id not in lookup:
                lookup[j.voucher_id] = j
            elif j.confirmed and j.voucher_id in lookup:
                # 取风险级别更高的
                if j.risk_level == "高":
                    lookup[j.voucher_id] = j
    return lookup


def _build_hit_lookup(rule_results: list["RuleResult"]) -> dict[str, list["RuleHit"]]:
    lookup: dict = {}
    for rr in rule_results:
        for hit in rr.hits:
            lookup.setdefault(hit.voucher_id, []).append(hit)
    return lookup


def _write_sample_sheet(wb, df, confirmed_vids, hit_lookup, judgment_lookup):
    ws = wb.active
    ws.title = "样本清单"
    voucher_rows_map = {
        str(vid): grp
        for vid, grp in df.groupby(df["凭证编号"].astype(str), sort=False)
    }

    headers = [c for c, _ in SAMPLE_COLS]
    widths = [w for _, w in SAMPLE_COLS]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row_num = 2
    for seq, vid in enumerate(confirmed_vids, 1):
        voucher_rows = voucher_rows_map.get(vid)
        if voucher_rows is None:
            voucher_rows = df.iloc[0:0]
        hits = hit_lookup.get(vid, [])
        judgment = judgment_lookup.get(vid)

        rule_types = " | ".join(dict.fromkeys(h.rule_type for h in hits))
        evidences = " | ".join(dict.fromkeys(h.evidence for h in hits))
        group_ids = " | ".join(dict.fromkeys(h.group_id for h in hits if h.group_id))
        related_vouchers = " | ".join(dict.fromkeys(
            related for h in hits for related in h.related_voucher_ids
        ))
        relation_evidences = " | ".join(dict.fromkeys(
            h.relation_evidence for h in hits if h.relation_evidence
        ))
        year = hits[0].year if hits and hits[0].year else (
            int(voucher_rows["_year"].iloc[0]) if "_year" in voucher_rows.columns and not voucher_rows.empty else ""
        )

        for _, row in voucher_rows.iterrows():
            amt = row.get("凭证货币价值")
            values = [
                seq,
                vid,
                group_ids,
                related_vouchers,
                year,
                row.get("过账日期"),
                row.get("凭证类型"),
                row.get("总账科目"),
                row.get("总账科目：长文本"),
                row.get("借/贷标识"),
                amt,
                str(row.get("文本", ""))[:80],
                row.get("供应商科目：名称 1"),
                row.get("客户科目：姓名 1"),
                row.get("用户名"),
                rule_types,
                evidences[:100],
                relation_evidences[:200],
                judgment.risk_level if judgment else "",
                judgment.reason if judgment else "",
                judgment.audit_procedures if judgment else "",
                "",  # 核查结论留白
            ]

            fill = HIGH_FILL if (judgment and judgment.risk_level == "高") else (
                MED_FILL if judgment else None
            )

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP
                if fill:
                    cell.fill = fill

            row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(SAMPLE_COLS))}1"


def _write_manual_final_sheet(wb, df, manual_final_samples):
    ws = wb.create_sheet("人工直入样本")
    headers = [
        "疑点群体", "来源模块", "来源视图", "标签", "入样理由",
        "凭证编号", "年份", "过账日期", "凭证类型", "总账科目", "总账科目名称",
        "借/贷", "金额", "文本", "供应商", "客户", "用户名",
    ]
    widths = [28, 14, 20, 24, 36, 14, 8, 12, 10, 12, 22, 8, 16, 34, 22, 22, 12]

    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_num = 2
    for group in manual_final_samples:
        voucher_ids = {str(vid) for vid in group.get("voucher_ids", [])}
        if not voucher_ids:
            continue
        rows = df[df["凭证编号"].astype(str).isin(voucher_ids)]
        for _, row in rows.iterrows():
            values = [
                group.get("title", ""),
                group.get("source_module", ""),
                group.get("source_view", ""),
                " | ".join(group.get("tags", [])),
                group.get("reason", ""),
                row.get("凭证编号"),
                int(row.get("_year")) if "_year" in row and row.get("_year") == row.get("_year") else "",
                row.get("过账日期"),
                row.get("凭证类型"),
                row.get("总账科目"),
                row.get("总账科目：长文本"),
                row.get("借/贷标识"),
                row.get("凭证货币价值"),
                str(row.get("文本", ""))[:120],
                row.get("供应商科目：名称 1"),
                row.get("客户科目：姓名 1"),
                row.get("用户名"),
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = WRAP
                cell.fill = MED_FILL
            row_num += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _write_stats_sheet(wb, rule_results, llm_judgments):
    ws = wb.create_sheet("规则统计")
    headers = ["规则名称", "命中凭证数", "LLM确认数", "确认率", "高风险", "中风险"]
    widths = [22, 14, 14, 10, 10, 10]

    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    total_hits = total_confirmed = total_high = total_med = 0

    for row_idx, rr in enumerate(rule_results, 2):
        voucher_count = len({h.voucher_id for h in rr.hits})
        judgments = llm_judgments.get(rr.rule_name, [])
        confirmed = len(judgments)
        high = sum(1 for j in judgments if j.risk_level == "高")
        med = sum(1 for j in judgments if j.risk_level == "中")
        rate = f"{confirmed / voucher_count:.0%}" if voucher_count > 0 else "N/A"

        for col_idx, val in enumerate([rr.rule_name, voucher_count, confirmed, rate, high, med], 1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = THIN_BORDER

        total_hits += voucher_count
        total_confirmed += confirmed
        total_high += high
        total_med += med

    total_row = len(rule_results) + 2
    total_rate = f"{total_confirmed / total_hits:.0%}" if total_hits > 0 else "N/A"
    for col_idx, val in enumerate(["合计", total_hits, total_confirmed, total_rate, total_high, total_med], 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"


def _write_rules_sheet(wb, rule_results):
    """规则命中明细：每条 RuleHit 一行，供追溯。"""
    ws = wb.create_sheet("命中明细")
    headers = ["规则名称", "规则类型", "凭证编号", "组合ID", "关联凭证", "优先级", "触发证据", "组合证据"]
    widths = [18, 24, 14, 24, 28, 8, 60, 70]

    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_num = 2
    for rr in rule_results:
        for hit in sorted(rr.hits, key=lambda h: h.priority, reverse=True):
            for col_idx, val in enumerate(
                [
                    rr.rule_name,
                    hit.rule_type,
                    hit.voucher_id,
                    hit.group_id or "",
                    " | ".join(hit.related_voucher_ids),
                    hit.priority,
                    hit.evidence,
                    hit.relation_evidence,
                ],
                1,
            ):
                ws.cell(row=row_num, column=col_idx, value=val).border = THIN_BORDER
            row_num += 1

    ws.freeze_panes = "A2"
