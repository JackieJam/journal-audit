"""
序时账审计抽样工具 — Excel 输出模块
生成样本清单和规则统计两个 Sheet
"""

from __future__ import annotations

from datetime import datetime
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    import pandas as pd

    from llm_analyzer import LLMJudgment
    from rules import RuleResult


# ============================================================
# 样式常量
# ============================================================

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HIGH_RISK_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
MEDIUM_RISK_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")

SAMPLE_COLUMNS = [
    ("凭证编号", 14),
    ("过账日期", 12),
    ("凭证类型", 10),
    ("总账科目", 14),
    ("总账科目名称", 22),
    ("借/贷标识", 10),
    ("凭证货币价值", 16),
    ("公司代码货币价值", 16),
    ("文本", 30),
    ("供应商编号", 14),
    ("供应商名称", 22),
    ("客户", 14),
    ("客户名称", 22),
    ("用户名", 14),
    ("规则类型", 20),
    ("触发证据", 30),
    ("LLM风险级别", 12),
    ("LLM判断理由", 30),
    ("建议核查程序", 30),
]


def _apply_header_style(ws, col_count: int):
    """给表头行设置样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _set_column_widths(ws, widths: list[int]):
    """设置列宽"""
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def generate_output(
    df: pd.DataFrame,
    rule_results: list[RuleResult],
    llm_judgments: dict[str, list[LLMJudgment]],
    output_path: str,
):
    """
    生成最终 Excel 报告。

    Args:
        df: 原始序时账 DataFrame
        rule_results: 规则引擎输出
        llm_judgments: LLM 核实结果 {rule_name: [LLMJudgment]}
        output_path: 输出文件路径
    """
    wb = Workbook()

    # ============================================================
    # Sheet1: 样本清单
    # ============================================================
    ws1 = wb.active
    ws1.title = "样本清单"

    # 写表头
    for col_idx, (col_name, _) in enumerate(SAMPLE_COLUMNS, 1):
        ws1.cell(row=1, column=col_idx, value=col_name)
    _apply_header_style(ws1, len(SAMPLE_COLUMNS))
    _set_column_widths(ws1, [w for _, w in SAMPLE_COLUMNS])

    # 构建 LLM 判断的 lookup：{voucher_id: (rule_name, LLMJudgment)}
    judgment_lookup: dict[float, tuple[str, LLMJudgment]] = {}
    for rule_name, judgments in llm_judgments.items():
        for j in judgments:
            judgment_lookup[j.voucher_id] = (rule_name, j)

    # 收集所有确认的凭证（去重）
    confirmed_vids: set[float] = set()
    hit_lookup: dict[float, list[RuleHit]] = {}

    for rule_result in rule_results:
        for hit in rule_result.hits:
            vid = hit.voucher_id
            # 保留 LLM 确认的，或者未进入 LLM 但优先级 ≥ 4 的
            if vid in judgment_lookup or hit.priority >= 4:
                confirmed_vids.add(vid)
                hit_lookup.setdefault(vid, []).append(hit)

    # 按优先级降序排列
    confirmed_list = sorted(
        confirmed_vids,
        key=lambda v: max(h.priority for h in hit_lookup.get(v, [])),
        reverse=True,
    )

    row_num = 2
    for vid in confirmed_list:
        voucher_rows = df[df["凭证编号"] == vid]
        hits = hit_lookup.get(vid, [])
        judgment = judgment_lookup.get(vid)

        for _, row in voucher_rows.iterrows():
            values = [
                row.get("凭证编号"),
                row.get("过账日期"),
                row.get("凭证类型"),
                str(int(row["总账科目"])) if isinstance(row.get("总账科目"), float) else row.get("总账科目"),
                row.get("总账科目：长文本"),
                row.get("借/贷标识"),
                row.get("凭证货币价值"),
                row.get("公司代码货币价值"),
                row.get("文本"),
                row.get("供应商编号"),
                row.get("供应商科目：名称 1"),
                row.get("客户"),
                row.get("客户科目：姓名 1"),
                row.get("用户名"),
                " | ".join(h.rule_type for h in hits) if hits else "",
                " | ".join(h.evidence for h in hits) if hits else "",
                judgment[1].risk_level if judgment else "",
                judgment[1].reason if judgment else "",
                judgment[1].audit_procedures if judgment else "",
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws1.cell(row=row_num, column=col_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGNMENT

                # 高风险行高亮
                if judgment and judgment[1].risk_level == "高":
                    cell.fill = HIGH_RISK_FILL
                elif judgment and judgment[1].risk_level == "中":
                    cell.fill = MEDIUM_RISK_FILL

            row_num += 1

    # 冻结首行
    ws1.freeze_panes = "A2"

    # ============================================================
    # Sheet2: 规则命中统计
    # ============================================================
    ws2 = wb.create_sheet("规则统计")

    stat_headers = ["规则名称", "规则命中数", "LLM确认数", "确认率"]
    for col_idx, header in enumerate(stat_headers, 1):
        ws2.cell(row=1, column=col_idx, value=header)
    _apply_header_style(ws2, len(stat_headers))

    stat_col_widths = [22, 14, 14, 12]
    _set_column_widths(ws2, stat_col_widths)

    for row_idx, rule_result in enumerate(rule_results, 2):
        hit_count = rule_result.count
        judgments = llm_judgments.get(rule_result.rule_name, [])
        confirmed_count = len(judgments)
        rate = f"{confirmed_count / hit_count:.0%}" if hit_count > 0 else "N/A"

        ws2.cell(row=row_idx, column=1, value=rule_result.rule_name).border = THIN_BORDER
        ws2.cell(row=row_idx, column=2, value=hit_count).border = THIN_BORDER
        ws2.cell(row=row_idx, column=3, value=confirmed_count).border = THIN_BORDER
        ws2.cell(row=row_idx, column=4, value=rate).border = THIN_BORDER

    # 汇总行
    total_row = len(rule_results) + 2
    total_hits = sum(r.count for r in rule_results)
    total_confirmed = sum(len(llm_judgments.get(r.rule_name, [])) for r in rule_results)
    total_rate = f"{total_confirmed / total_hits:.0%}" if total_hits > 0 else "N/A"

    ws2.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
    ws2.cell(row=total_row, column=1).border = THIN_BORDER
    ws2.cell(row=total_row, column=2, value=total_hits).font = Font(bold=True)
    ws2.cell(row=total_row, column=2).border = THIN_BORDER
    ws2.cell(row=total_row, column=3, value=total_confirmed).font = Font(bold=True)
    ws2.cell(row=total_row, column=3).border = THIN_BORDER
    ws2.cell(row=total_row, column=4, value=total_rate).font = Font(bold=True)
    ws2.cell(row=total_row, column=4).border = THIN_BORDER

    ws2.freeze_panes = "A2"

    # ============================================================
    # 保存
    # ============================================================
    wb.save(output_path)
    print(f"\n✓ 报告已生成: {output_path}")
    print(f"  样本条数: {row_num - 2} 行（含凭证全部行项目）")
    print(f"  涉及凭证: {len(confirmed_list)} 个")
    print(f"  规则命中: {total_hits} 次，LLM 确认: {total_confirmed} 次")
